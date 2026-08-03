"""Export Excel du tableau d'extraction (Feuil2 de `donnes_ref_v2.md`, ligne « ajouter extraction
tableau dans excel »)."""
from __future__ import annotations

import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.extraction.excel_export import (
    SHEET_TITLE,
    build_extraction_workbook,
    extraction_workbook_filename,
)
from app.extraction.extraction_schema import load_extraction_schema
from app.extraction.pipeline import ensure_results_initialized
from app.store.db import session_scope
from app.store.repository import (
    create_dossier,
    get_dossier,
    get_extraction_result_by_field,
    set_extraction_result,
)


def _dossier_with_results(original_filename: str = "DCE Marly.zip") -> str:
    with session_scope() as s:
        dossier = create_dossier(s, original_filename=original_filename)
        dossier_id = dossier.id
        ensure_results_initialized(s, dossier_id)
    return dossier_id


def _sheet(dossier_id: str):
    with session_scope() as s:
        dossier = get_dossier(s, dossier_id)
        content = build_extraction_workbook(s, dossier)
    return load_workbook(io.BytesIO(content))[SHEET_TITLE]


def _set(dossier_id: str, field_id: str, **overrides) -> None:
    defaults = dict(
        match_layer="file",
        value="",
        confidence=None,
        justification="",
        citation="",
        sources=[],
        cross_check_status=None,
        model_name="m",
        model_version="v",
        error=None,
    )
    defaults.update(overrides)
    with session_scope() as s:
        result = get_extraction_result_by_field(s, dossier_id, field_id)
        set_extraction_result(s, result, **defaults)


def test_workbook_contains_every_schema_field_including_the_missing_ones(isolated_workspace):
    """L'absence d'une donnée de souscription est une information : un champ non trouvé doit
    apparaître avec une valeur vide, pas disparaître du classeur."""
    dossier_id = _dossier_with_results()
    schema = load_extraction_schema()

    sheet = _sheet(dossier_id)
    libelles = {row[1] for row in sheet.iter_rows(min_row=2, values_only=True) if row[1]}

    for f in schema.fields:
        assert f.libelle in libelles, f"{f.id} absent de l'export Excel"


def test_workbook_groups_fields_under_section_banners_in_schema_order(isolated_workspace):
    dossier_id = _dossier_with_results()
    schema = load_extraction_schema()

    sheet = _sheet(dossier_id)
    # Un bandeau de section n'a que sa première cellule remplie.
    banners = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] and row[1] is None]

    assert banners == [s.libelle for s in schema.sections]


def test_workbook_reports_final_value_sources_and_cross_check(isolated_workspace):
    dossier_id = _dossier_with_results()
    _set(
        dossier_id,
        "montants_totaux_ht",
        value="12 000 000 €",
        confidence=0.92,
        justification="Montant lu à l'article 3.",
        citation="Montant des travaux : 12 000 000 € HT",
        sources=[{"document_id": "d1", "filename": "RC.pdf", "value": "12 000 000 €", "confidence": 0.92}],
        cross_check_status="coherent",
    )

    sheet = _sheet(dossier_id)
    row = next(r for r in sheet.iter_rows(min_row=2, values_only=True) if r[1] == "Montants totaux HT")

    assert row[2] == "12 000 000 €"
    assert row[5] == "Cohérent"
    assert row[6] == 0.92
    assert "RC.pdf" in row[10]


def test_workbook_flags_an_incoherent_cross_check(isolated_workspace):
    """L'un des deux seuls signaux visuels du classeur, parce qu'il appelle un arbitrage humain."""
    dossier_id = _dossier_with_results()
    _set(dossier_id, "montants_totaux_ht", value="12 000 000 €", cross_check_status="incoherent")

    sheet = _sheet(dossier_id)
    row_index = next(r[0].row for r in sheet.iter_rows(min_row=2) if r[1].value == "Montants totaux HT")

    assert sheet.cell(row=row_index, column=6).value == "INCOHÉRENT"
    assert sheet.cell(row=row_index, column=6).fill.fgColor.rgb.endswith("FCE4E4")


def test_workbook_escapes_nothing_but_keeps_multiline_sources(isolated_workspace):
    """Contrairement au Markdown, une cellule Excel accepte les retours à la ligne : plusieurs
    fichiers sources restent lisibles empilés dans la même cellule."""
    dossier_id = _dossier_with_results()
    _set(
        dossier_id,
        "montants_totaux_ht",
        value="12 M€",
        sources=[
            {"document_id": "d1", "filename": "RC.pdf", "value": "12 M€", "confidence": 0.9},
            {"document_id": "d2", "filename": "CCAP.pdf", "value": "11 M€", "confidence": 0.8},
        ],
    )

    sheet = _sheet(dossier_id)
    row = next(r for r in sheet.iter_rows(min_row=2, values_only=True) if r[1] == "Montants totaux HT")

    assert row[10] == "RC.pdf\nCCAP.pdf"


def test_filename_is_derived_from_the_original_dossier_name():
    class _FakeDossier:
        id = "abc123"
        original_filename = "DCE Marly / conservatoire.zip"

    assert extraction_workbook_filename(_FakeDossier()) == "extraction_DCE Marly _ conservatoire.xlsx"


def test_export_endpoint_serves_an_xlsx(isolated_workspace):
    from app.main import app

    dossier_id = _dossier_with_results()
    client = TestClient(app)

    response = client.get(f"/api/dossiers/{dossier_id}/extraction/export.xlsx")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in response.headers["content-disposition"]
    assert load_workbook(io.BytesIO(response.content)).sheetnames == [SHEET_TITLE]


def test_export_endpoint_404s_on_unknown_dossier(isolated_workspace):
    from app.main import app

    client = TestClient(app)
    assert client.get("/api/dossiers/inconnu/extraction/export.xlsx").status_code == 404
