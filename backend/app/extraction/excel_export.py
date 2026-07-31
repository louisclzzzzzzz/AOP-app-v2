"""Export Excel du tableau d'extraction (Feuil2 de `refs/donnes_ref_v2.md`, ligne « ajouter
extraction tableau dans excel »).

L'expert métier travaille la Feuil2 dans un classeur : lui rendre le résultat sous la même forme
lui évite de recopier à la main les 50 valeurs depuis l'écran de validation ou le rapport Markdown.
Le classeur reprend donc l'ordre et le regroupement par section d'`extraction_schema.yaml`, avec
une colonne par information nécessaire à la vérification (valeur, sources, preuve, confiance).

Écrit dans un buffer mémoire et non sur disque : contrairement aux rapports JSON/Markdown, ce
fichier n'a pas à être figé au checkpoint — il est régénéré à chaque téléchargement à partir de
l'état courant en base, donc il reflète toujours les dernières corrections manuelles.
"""
from __future__ import annotations

import io
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.extraction.extraction_schema import load_extraction_schema
from app.store.models import Dossier
from app.store.repository import list_documents, list_extraction_results

SHEET_TITLE = "Extraction"

# (en-tête, largeur de colonne). Les largeurs sont fixées à la main : `openpyxl` ne sait pas
# auto-ajuster, et une colonne « Citation » à largeur par défaut rend le classeur illisible.
_COLUMNS: list[tuple[str, int]] = [
    ("Section", 28),
    ("Donnée", 38),
    ("Valeur", 52),
    ("Résultat attendu", 46),
    ("Statut", 14),
    ("Recoupement", 20),
    ("Confiance", 11),
    ("Corrigé manuellement", 12),
    ("Justification", 52),
    ("Citation", 60),
    ("Fichiers sources", 44),
]

_CROSS_CHECK_LABELS = {
    "coherent": "Cohérent",
    "incoherent": "INCOHÉRENT",
    "single_source": "Source unique",
    "not_applicable": "",
}

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
_INCOHERENT_FILL = PatternFill("solid", fgColor="FCE4E4")
_MISSING_FILL = PatternFill("solid", fgColor="F2F2F2")


def _source_labels(result, documents_by_id: dict[str, object]) -> str:
    if not result.proposed_sources_json:
        return ""
    sources = json.loads(result.proposed_sources_json)
    labels: list[str] = []
    for source in sources:
        doc = documents_by_id.get(source.get("document_id"))
        # Le chemin relatif situe le document dans l'arborescence réorganisée ; on retombe sur le
        # nom de fichier seul quand le document n'existe plus (dossier rejoué, document retiré).
        labels.append(getattr(doc, "relative_path", None) or source.get("filename") or "?")
    return "\n".join(labels)


def build_extraction_workbook(session: Session, dossier: Dossier) -> bytes:
    """Classeur .xlsx du tableau d'extraction, à partir de l'état courant en base (propositions du
    modèle et corrections humaines confondues — c'est `final_value` qui fait foi)."""
    schema = load_extraction_schema()
    results_by_field = {r.field_id: r for r in list_extraction_results(session, dossier.id)}
    documents_by_id = {d.id: d for d in list_documents(session, dossier.id)}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_TITLE

    sheet.append([header for header, _ in _COLUMNS])
    for index, (_, width) in enumerate(_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    current_section: str | None = None
    for extraction_field in schema.fields:
        # Bandeau de section : le classeur reprend le découpage thématique du schéma, sinon les
        # 50 lignes de la Feuil2 se lisent comme une liste indifférenciée.
        if extraction_field.section != current_section:
            current_section = extraction_field.section
            sheet.append([schema.section_label(current_section)])
            banner = sheet.cell(row=sheet.max_row, column=1)
            banner.font = Font(bold=True)
            banner.fill = _SECTION_FILL

        result = results_by_field.get(extraction_field.id)
        value = (result.final_value if result else None) or ""
        sheet.append(
            [
                schema.section_label(extraction_field.section),
                extraction_field.libelle,
                value,
                extraction_field.resultat_attendu or "",
                (result.status if result else "") or "",
                _CROSS_CHECK_LABELS.get(result.cross_check_status if result else None, "") or "",
                result.proposed_confidence if result else None,
                "Oui" if result and result.is_manually_corrected else "",
                (result.proposed_justification if result else None) or "",
                (result.proposed_citation if result else None) or "",
                _source_labels(result, documents_by_id) if result else "",
            ]
        )

        row = sheet[sheet.max_row]
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        # Deux signaux visuels, les seuls qui déclenchent une action de l'expert : une valeur
        # absente (à aller chercher à la main) et un recoupement incohérent (à arbitrer).
        if not value:
            row[2].fill = _MISSING_FILL
        if result and result.cross_check_status == "incoherent":
            row[5].fill = _INCOHERENT_FILL

    sheet.freeze_panes = "C2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def extraction_workbook_filename(dossier: Dossier) -> str:
    """Nom de fichier proposé au téléchargement, dérivé du nom du dossier d'origine."""
    stem = (dossier.original_filename or dossier.id).rsplit(".", 1)[0]
    safe = "".join(c if c.isalnum() or c in " -_" else "_" for c in stem).strip() or dossier.id
    return f"extraction_{safe}.xlsx"
